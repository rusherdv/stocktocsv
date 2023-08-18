import mysql.connector
import customtkinter
from tkinter import *
import os
from time import sleep
import locale
import random
import datetime
from datetime import datetime
from os import mkdir
import xlsxwriter
import openpyxl

customtkinter.set_appearance_mode("dark")
customtkinter.set_default_color_theme("dark-blue")

root = customtkinter.CTk()
root.geometry("700x900")

root.title("Sistema de Registro de Ventas")

def main():
    
    today = datetime.today().strftime('%d-%m-%y')
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
    currentYear = str(datetime.now().year)
    currentMonth = str(datetime.now().strftime("%B"))

    datos = [["Planilla Diaria", today, "", "" , ""],["Producto", "SKU", "Cantidad", "Metodo de Pago" , "Precio"]]
    
    mydb = mysql.connector.connect(
        host="localhost",
        username="root",
        password="",
        database="stockdb"
    )
    
    mycursor = mydb.cursor()

    def searchDatabase():
        id = entry.get()
        cantidad = entry1.get()
        precio = entry2.get()
        if (id == ''):
            error("Complete el codigo de barras")
        elif (cantidad == ""):
            error("Complete la cantidad")
        elif (precio == ''):
            error("Complete el precio")
        elif (combobox.get() == 'Elige una opcion'):
            error("Elige un metodo de pago")
        else:
            mycursor.execute("SELECT * from stock WHERE sku = %s", (id,))
            myresult = mycursor.fetchall()
            if (str(myresult) == '[]'):
                error("No se encontro el producto")
            else:
                for x in myresult:
                    sku = x[0]
                    nombre = x[1]
                    checkFolder()
                    checkFile(nombre, sku)

    def savingArray(nombre, sku):
        cant = entry1.get()
        payMethod = combobox.get()
        price = entry2.get()

        if (switch_1.get() == "on"):
            datos.append([nombre, sku, "x" + cant, payMethod, "$" + str(int(price) * int(cant))])
        elif (switch_1.get() == "off"):
            datos.append([nombre, sku, "x" +  cant, payMethod, "$" +  str(int(price) * int(cant))]) 
            datos.append(["", "", "", "" , ""])


    def createFolder():
        os.makedirs("Planillas/" + currentYear + "/" + currentMonth)
                
    def createFile():
        print("Creando archivo")
        archivo=xlsxwriter.Workbook("Planillas/" + currentYear + "/" + currentMonth + "/" + "Planilla - " + str(datetime.now().day) + '.xlsx')
        format = archivo.add_format()
        format.set_font_size(20)
        hoja=archivo.add_worksheet()

        for j in range(len(datos)):
            for i in range(len(datos[j])):
                hoja.write(j,i,datos[j][i])
        archivo.close()
        guardado()

    def guardado():
        entry.delete('0', 'end')
        entry1.delete('0', 'end')
        entry2.delete('0', 'end')
        combobox.set('Elige una opcion')
        entry.focus()
        label5.configure(text="Guardado", text_color="green")
        label5.pack(pady=10, padx=0)

        root.after(3000, hideSave)

    def hideSave():
        label5.pack_forget()

    def error(problema):
        label5.configure(text=problema, text_color="red")
        label5.pack(pady=10, padx=0)

    def checkFile(nombre, sku):
        archivo = "Planillas/" + currentYear + "/" + currentMonth + "/" + "Planilla - " + str(datetime.now().day) + '.xlsx'

        if (os.path.isfile(archivo) == False):
            savingArray(nombre, sku)
            createFile()
        elif (os.path.isfile(archivo) == True):
            archivo="Planillas/" + currentYear + "/" + currentMonth + "/" + "Planilla - " + str(datetime.now().day) + '.xlsx'
            
            wb_obj = openpyxl.load_workbook(archivo)
            sheet_obj = wb_obj.active
            m_row = sheet_obj.max_row

            for i in range(1, m_row + 1):
                if (switch_1.get() == "on"):
                    proxRow = "A" + str(m_row + 1)
                    sheet_obj[proxRow].value = nombre
                    proxRow = "B" + str(m_row + 1)
                    sheet_obj[proxRow].value = sku
                    proxRow = "C" + str(m_row + 1)
                    sheet_obj[proxRow].value = "x" + entry1.get()
                    proxRow = "D" + str(m_row + 1)
                    sheet_obj[proxRow].value = combobox.get()
                    proxRow = "E" + str(m_row + 1)
                    cantidad = int(entry1.get())
                    precio = int(entry2.get())
                    sheet_obj[proxRow].value = "$" + str(cantidad * precio)

                else:
                    proxRow = "A" + str(m_row + 2)
                    sheet_obj[proxRow].value = nombre
                    proxRow = "B" + str(m_row + 2)
                    sheet_obj[proxRow].value = sku
                    proxRow = "C" + str(m_row + 2)
                    sheet_obj[proxRow].value = "x" + entry1.get()
                    proxRow = "D" + str(m_row + 2)
                    sheet_obj[proxRow].value = combobox.get()
                    proxRow = "E" + str(m_row + 2)
                    cantidad = int(entry1.get())
                    precio = int(entry2.get())
                    sheet_obj[proxRow].value = "$" + str(cantidad * precio)
                    
                    
                wb_obj.save(archivo)
                guardado()

            

    def checkFolder():
        if (os.path.isdir("Planillas/" + currentYear + "/" + currentMonth)):
            print('')
        else:
            createFolder()

    frame = customtkinter.CTkFrame(master=root)
    frame.pack(pady=25, padx=30, fill="both", expand=True)

    label = customtkinter.CTkLabel(master=frame, text="Sistema de Registro de Ventas", font=('Montserrat',30,'bold'))
    label.pack(pady=35, padx=10)

    label3 = customtkinter.CTkLabel(master=frame, text="Inserte codigo de barras: ", font=('Montserrat',14,'bold'))
    label3.pack(pady=10, padx=0)

    entry = customtkinter.CTkEntry(master=frame, placeholder_text="Codigo")
    entry.pack(pady=10, padx=10)
    
    label3 = customtkinter.CTkLabel(master=frame, text="Inserte la cantidad: ", font=('Montserrat',14,'bold'))
    label3.pack(pady=10, padx=0)

    entry1 = customtkinter.CTkEntry(master=frame, placeholder_text="Cantidad")
    entry1.insert(0, "1")
    entry1.pack(pady=10, padx=10)

    label2 = customtkinter.CTkLabel(master=frame, text="Eliga el metodo de pago: ", font=('Montserrat',14,'bold'))
    label2.pack(pady=10, padx=0)

    optionmenu_var = customtkinter.StringVar(value="")
    combobox = customtkinter.CTkOptionMenu(master=frame,values=["Efectivo","Tarjeta", "MercadoPago" ,"PedidosYa | Efectivo", "PedidosYa | Online", "Rappi"],variable=optionmenu_var)
    combobox.pack(padx=20, pady=10)
    combobox.set("Elige una opcion")
    combobox.pack(padx=10, pady=10)

    label4 = customtkinter.CTkLabel(master=frame, text="Inserte precio del producto: ", font=('Montserrat',14,'bold'))
    label4.pack(pady=10, padx=0)

    entry2 = customtkinter.CTkEntry(master=frame, placeholder_text="Precio")
    entry2.pack(pady=10, padx=10)

    switch_var = customtkinter.StringVar(value="off")
    switch_1 = customtkinter.CTkSwitch(master=frame, text="Misma Compra", variable=switch_var, onvalue="on", offvalue="off")
    switch_1.pack(padx=20, pady=10)
        
    button = customtkinter.CTkButton(master=frame, text="Cargar", command=searchDatabase)
    button.pack(pady=20, padx=10)

    label5 = customtkinter.CTkLabel(master=frame, text="Guardado", font=('Montserrat',14, 'bold'), text_color="green")

    root.mainloop()

main()